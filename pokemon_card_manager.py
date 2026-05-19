#!/usr/bin/env python3
import base64
import hashlib
import json
import os
import re
import shutil
import socket
import sys
import threading
import time as _time
import tkinter as tk
import urllib.error
import urllib.request
import webbrowser
from http.server import HTTPServer, BaseHTTPRequestHandler
from tkinter import ttk, messagebox, simpledialog, filedialog
from datetime import date
from urllib.parse import parse_qs

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import CellIsRule
from openpyxl.worksheet.datavalidation import DataValidation
from PIL import Image, ImageTk

VERSION = "1.2.0"
GITHUB_RAW_URL = "https://raw.githubusercontent.com/Werizu/PokemonCardManager/main/pokemon_card_manager.py"

BASE_DIR = os.path.join(os.path.expanduser("~"), "Pokemon-Sammlung")
EXCEL_PATH = os.path.join(BASE_DIR, "Pokemon-Inventar.xlsx")
KARTEN_DIR = os.path.join(BASE_DIR, "Karten")
CONFIG_PATH = os.path.join(BASE_DIR, ".app_config.json")

SETS = [
    "151", "Ancient Origins", "Aquapolis", "Arceus", "Ascended Heroes",
    "Astral Radiance", "Base Set", "Base Set 2", "Battle Styles",
    "Black & White", "Black Bolt", "Boundaries Crossed", "BREAKpoint",
    "BREAKthrough", "Brilliant Stars", "Burning Shadows", "Call of Legends",
    "Celebrations", "Celestial Storm", "Champion's Path", "Chilling Reign",
    "Cosmic Eclipse", "Crimson Invasion", "Crown Zenith", "Dark Explorers",
    "Darkness Ablaze", "Destined Rivals", "Detective Pikachu",
    "Diamond & Pearl", "Double Crisis", "Dragon Majesty", "Dragon Vault",
    "Dragons Exalted", "Emerging Powers", "Evolutions", "Evolving Skies",
    "EX Crystal Guardians", "EX Delta Species", "EX Deoxys", "EX Dragon",
    "EX Dragon Frontiers", "EX Emerald", "EX FireRed & LeafGreen",
    "EX Hidden Legends", "EX Holon Phantoms", "EX Legend Maker",
    "EX Power Keepers", "EX Ruby & Sapphire", "EX Sandstorm",
    "EX Team Magma vs Team Aqua", "EX Team Rocket Returns",
    "EX Unseen Forces", "Expedition Base Set", "Fates Collide", "Flashfire",
    "Forbidden Light", "Fossil", "Furious Fists", "Fusion Strike",
    "Generations", "Great Encounters", "Guardians Rising", "Gym Challenge",
    "Gym Heroes", "HeartGold & SoulSilver", "Hidden Fates", "HS Triumphant",
    "HS Undaunted", "HS Unleashed", "Journey Together", "Jungle",
    "Legendary Collection", "Legendary Treasures", "Legends Awakened",
    "Lost Origin", "Lost Thunder", "Majestic Dawn", "Mega Evolution",
    "Mysterious Treasures", "Neo Destiny", "Neo Discovery", "Neo Genesis",
    "Neo Revelation", "Next Destinies", "Noble Victories", "Obsidian Flames",
    "Other", "Perfect Order", "Paldea Evolved", "Paldean Fates",
    "Paradox Rift", "Phantasmal Flames", "Phantom Forces", "Plasma Blast",
    "Plasma Freeze", "Plasma Storm", "Platinum", "Pokemon GO", "Primal Clash",
    "Prismatic Evolutions", "Promo", "Rebel Clash", "Rising Rivals",
    "Roaring Skies", "Scarlet & Violet", "Secret Wonders", "Shining Fates",
    "Shining Legends", "Shrouded Fable", "Silver Tempest", "Skyridge",
    "Steam Siege", "Stellar Crown", "Stormfront", "Sun & Moon",
    "Supreme Victors", "Surging Sparks", "Sword & Shield", "Team Rocket",
    "Team Up", "Temporal Forces", "Twilight Masquerade", "Ultra Prism",
    "Unbroken Bonds", "Unified Minds", "Vivid Voltage", "White Flare", "XY",
]

LANGUAGES = ["EN", "DE", "JP", "KR", "FR", "IT", "ES", "CN", "PT"]
CONDITIONS = ["Mint", "Near Mint", "Excellent", "Good", "Light Played", "Played", "Poor"]
SOURCES = ["Cardmarket", "eBay", "Convention", "Trade", "Booster", "Other"]

# Col mapping: 1=ID, 2=Name, 3=Set, 4=Code, 5=Lang, 6=Condition,
# 7=Grading Status, 8=Grading Service, 9=Grade, 10=Cert#,
# 11=Purchase Price, 12=Purchase Date, 13=Source,
# 14=Quantity, 15=Status

GRADE_ALIASES = {
    "P10": "Pristine 10", "P 10": "Pristine 10",
    "B10": "Black Label 10", "B 10": "Black Label 10",
}

# eBay API
EBAY_CONFIG_PATH = os.path.join(BASE_DIR, ".ebay_config.json")
EBAY_SCOPES = ("https://api.ebay.com/oauth/api_scope/sell.inventory"
               " https://api.ebay.com/oauth/api_scope/sell.account")

CONDITION_TO_EBAY_CARD = {
    "Mint": "400010", "Near Mint": "400010",
    "Excellent": "400015", "Good": "400016",
    "Light Played": "400016", "Played": "400017", "Poor": "400017",
}

GRADING_SERVICE_TO_EBAY = {
    "PSA": "275010", "BCCG": "275011", "BVG": "275012", "BGS": "275013",
    "CGC": "275015", "SGC": "275016", "KSA": "275017", "GMA": "275018",
    "HGA": "275019", "ISA": "2750110", "PCA": "2750111", "GSG": "2750112",
    "PGS": "2750113", "MNT": "2750114", "TAG": "2750115", "Rare": "2750116",
    "RCG": "2750117", "PCG": "2750118", "Ace": "2750119", "CGA": "2750120",
    "TCG": "2750121", "ARK": "2750122", "AGS": "2750124", "DSG": "2750125",
}

GRADE_TO_EBAY = {
    "10": "275020", "9.5": "275021", "9": "275022", "8.5": "275023",
    "8": "275024", "7.5": "275025", "7": "275026", "6.5": "275027",
    "6": "275028", "5.5": "275029", "5": "2750210", "4.5": "2750211",
    "4": "2750212", "3.5": "2750213", "3": "2750214", "2.5": "2750215",
    "2": "2750216", "1.5": "2750217", "1": "2750218",
}


def parse_grade(grade_str):
    upper = grade_str.upper().strip()
    if upper in GRADE_ALIASES:
        return GRADE_ALIASES[upper]
    return float(grade_str)


def _match_set_name(slug):
    set_name = slug.replace("-", " ").replace(" and ", " & ")
    best_set = None
    best_score = 0
    set_lower = set_name.lower()
    for s in SETS:
        if s.lower() == set_lower:
            return s
        words_match = sum(1 for w in s.lower().split() if w in set_lower)
        if words_match > best_score:
            best_score = words_match
            best_set = s
    return best_set or set_name


def parse_cardmarket_url(url):
    url = url.strip()

    # Singles: /Products/Singles/{SetSlug}/{CardSlug}
    m_singles = re.search(
        r"cardmarket\.com/\w+/Pokemon/Products/Singles/([^/]+)/([^/?#]+)", url
    )
    if m_singles:
        set_slug = m_singles.group(1)
        card_slug = m_singles.group(2)
        parts = card_slug.split("-")
        card_code = ""
        known_variants = {"V1", "V2", "V3", "V4", "RH", "FA", "SAR", "SIR", "IR",
                          "AR", "HR", "SR", "UR", "RR", "GG", "TG", "GR"}
        if parts and re.match(r'^[A-Z]{2,5}\w*\d+$', parts[-1]):
            card_code = parts.pop()
        elif parts and parts[-1].isdigit():
            card_code = parts.pop()
        if parts and parts[-1] in known_variants:
            parts.pop()
        elif len(parts) >= 2 and f"{parts[-2]}-{parts[-1]}" == "Alt-Art":
            parts.pop()
            parts.pop()
        return {
            "name": " ".join(parts),
            "set": _match_set_name(set_slug),
            "number": card_code,
        }

    # Sealed products: /Products/{Type}/{ProductSlug}  (Boosters, Booster-Boxes, etc.)
    m_sealed = re.search(
        r"cardmarket\.com/\w+/Pokemon/Products/([^/]+)/([^/?#]+)", url
    )
    if m_sealed:
        product_type = m_sealed.group(1)
        if product_type == "Singles":
            return None
        product_slug = m_sealed.group(2)
        parts = product_slug.split("-")

        product_labels = {"Booster", "Bundle", "Display", "Box", "Elite", "Trainer",
                          "Case", "Pack", "Tin", "Mini", "Collection", "Premium",
                          "Pin", "Set", "10"}
        name_parts = []
        set_parts = []
        hit_label = False
        for p in parts:
            if p in product_labels:
                hit_label = True
            if hit_label:
                name_parts.append(p)
            else:
                set_parts.append(p)

        if not set_parts:
            set_parts = parts
            name_parts = parts

        set_slug = "-".join(set_parts)
        product_name = " ".join(name_parts) if name_parts else " ".join(parts)

        return {
            "name": product_name,
            "set": _match_set_name(set_slug),
            "number": "",
        }

    return None


# ============================================================
# EXCEL OPERATIONS
# ============================================================

def get_next_id():
    wb = load_workbook(EXCEL_PATH)
    ws = wb["Inventar"]
    max_id = 0
    for row in range(2, ws.max_row + 1):
        val = ws.cell(row=row, column=1).value
        if val is not None:
            try:
                max_id = max(max_id, int(val))
            except (ValueError, TypeError):
                pass
    wb.close()
    return max_id + 1


def save_card(name, set_name, card_number, language, condition, price, source,
              quantity=1, grading_status="Not Graded", grading_service="",
              grade=None, cert_number="", url=""):
    card_id = get_next_id()

    safe_set = set_name.replace(" ", "-").replace("&", "and").replace("'", "")
    safe_name = name.replace(" ", "-")
    folder_name = f"{card_id:03d}-{safe_name}"
    folder_path = os.path.join(KARTEN_DIR, safe_set, folder_name)
    os.makedirs(folder_path, exist_ok=True)

    wb = load_workbook(EXCEL_PATH)
    ws = wb["Inventar"]

    new_row = 2
    for row in range(2, 202):
        if ws.cell(row=row, column=1).value is None:
            new_row = row
            break

    ws.cell(row=new_row, column=1, value=card_id)
    ws.cell(row=new_row, column=2, value=name)
    ws.cell(row=new_row, column=3, value=set_name)
    ws.cell(row=new_row, column=4, value=card_number)
    ws.cell(row=new_row, column=5, value=language)
    ws.cell(row=new_row, column=6, value=condition)
    ws.cell(row=new_row, column=7, value=grading_status)
    if grading_service:
        ws.cell(row=new_row, column=8, value=grading_service)
    if grade is not None:
        ws.cell(row=new_row, column=9, value=grade)
    if cert_number:
        ws.cell(row=new_row, column=10, value=cert_number)
    ws.cell(row=new_row, column=11, value=price)
    ws.cell(row=new_row, column=11).number_format = '#,##0.00 "EUR"'
    ws.cell(row=new_row, column=12, value=date.today())
    ws.cell(row=new_row, column=12).number_format = "YYYY-MM-DD"
    ws.cell(row=new_row, column=13, value=source)
    ws.cell(row=new_row, column=14, value=quantity)
    ws.cell(row=new_row, column=15, value="Collection")

    wb.save(EXCEL_PATH)
    wb.close()
    if url:
        _save_url(card_id, url)
    return card_id


def load_all_cards():
    wb = load_workbook(EXCEL_PATH)
    ws = wb["Inventar"]
    cards = []
    for row in range(2, 202):
        card_id = ws.cell(row=row, column=1).value
        if card_id is None:
            continue
        cards.append({
            "row": row,
            "id": int(card_id),
            "name": ws.cell(row=row, column=2).value or "",
            "set": ws.cell(row=row, column=3).value or "",
            "number": ws.cell(row=row, column=4).value or "",
            "language": ws.cell(row=row, column=5).value or "",
            "price": ws.cell(row=row, column=11).value or 0,
            "quantity": ws.cell(row=row, column=14).value or 1,
            "status": ws.cell(row=row, column=15).value or "Collection",
            "grading_status": ws.cell(row=row, column=7).value or "Not Graded",
            "grading_service": ws.cell(row=row, column=8).value or "",
            "grade": ws.cell(row=row, column=9).value or "",
        })
    wb.close()
    return cards


def update_card_status(card_id, status):
    wb = load_workbook(EXCEL_PATH)
    ws = wb["Inventar"]
    for row in range(2, 202):
        if ws.cell(row=row, column=1).value == card_id:
            ws.cell(row=row, column=15, value=status)
            break
    wb.save(EXCEL_PATH)
    wb.close()


def mark_card_sold(card_id, sale_price, platform, fees, sell_qty):
    wb = load_workbook(EXCEL_PATH)
    ws_inv = wb["Inventar"]
    ws_sales = wb["Sales"]

    card_name = ""
    purchase_price = 0
    total_qty = 1
    inv_row = None
    for row in range(2, 202):
        if ws_inv.cell(row=row, column=1).value == card_id:
            inv_row = row
            card_name = ws_inv.cell(row=row, column=2).value or ""
            purchase_price = ws_inv.cell(row=row, column=11).value or 0
            total_qty = int(ws_inv.cell(row=row, column=14).value or 1)
            break

    if inv_row is None:
        wb.close()
        return

    remaining = total_qty - sell_qty
    if remaining <= 0:
        ws_inv.cell(row=inv_row, column=14, value=0)
        ws_inv.cell(row=inv_row, column=15, value="Sold")
    else:
        ws_inv.cell(row=inv_row, column=14, value=remaining)

    sale_row = 2
    for row in range(2, 202):
        if ws_sales.cell(row=row, column=1).value is None:
            sale_row = row
            break

    ws_sales.cell(row=sale_row, column=1, value=card_id)
    ws_sales.cell(row=sale_row, column=2, value=card_name)
    ws_sales.cell(row=sale_row, column=3, value=date.today())
    ws_sales.cell(row=sale_row, column=3).number_format = "YYYY-MM-DD"
    ws_sales.cell(row=sale_row, column=4, value=platform)
    ws_sales.cell(row=sale_row, column=5, value=sale_price)
    ws_sales.cell(row=sale_row, column=5).number_format = '#,##0.00 "EUR"'
    ws_sales.cell(row=sale_row, column=6, value=fees)
    ws_sales.cell(row=sale_row, column=6).number_format = '#,##0.00 "EUR"'
    r = sale_row
    ws_sales.cell(row=r, column=8).value = f'=IF(E{r}<>"",E{r}-F{r}-G{r},"")'
    ws_sales.cell(row=r, column=8).number_format = '#,##0.00 "EUR"'
    ws_sales.cell(row=r, column=9, value=purchase_price * sell_qty)
    ws_sales.cell(row=r, column=9).number_format = '#,##0.00 "EUR"'
    ws_sales.cell(row=r, column=10).value = f'=IF(AND(H{r}<>"",I{r}<>""),H{r}-I{r},"")'
    ws_sales.cell(row=r, column=10).number_format = '#,##0.00 "EUR"'

    wb.save(EXCEL_PATH)
    wb.close()


def submit_for_grading(card_id, service):
    wb = load_workbook(EXCEL_PATH)
    ws = wb["Inventar"]
    for row in range(2, 202):
        if ws.cell(row=row, column=1).value == card_id:
            ws.cell(row=row, column=7, value="Submitted")
            ws.cell(row=row, column=8, value=service)
            break
    wb.save(EXCEL_PATH)
    wb.close()


def grading_returned(card_id, grade, cert_number):
    wb = load_workbook(EXCEL_PATH)
    ws = wb["Inventar"]
    for row in range(2, 202):
        if ws.cell(row=row, column=1).value == card_id:
            ws.cell(row=row, column=7, value="Graded")
            ws.cell(row=row, column=9, value=grade)
            ws.cell(row=row, column=10, value=cert_number)
            break
    wb.save(EXCEL_PATH)
    wb.close()


def load_statistics():
    wb = load_workbook(EXCEL_PATH)
    ws_inv = wb["Inventar"]
    ws_sales = wb["Sales"]

    stats = {
        "total_items": 0,
        "in_collection": 0,
        "sold": 0,
        "at_grading": 0,
        "graded": 0,
        "total_spent": 0.0,
        "total_sale_revenue": 0.0,
        "total_fees": 0.0,
        "total_sale_profit": 0.0,
    }

    for row in range(2, 202):
        card_id = ws_inv.cell(row=row, column=1).value
        if card_id is None:
            continue
        quantity = ws_inv.cell(row=row, column=14).value or 1
        purchase = ws_inv.cell(row=row, column=11).value or 0
        stats["total_items"] += int(quantity)
        stats["total_spent"] += float(purchase) * int(quantity)

        status = ws_inv.cell(row=row, column=15).value or "Collection"
        grading_status = ws_inv.cell(row=row, column=7).value or "Not Graded"

        if status == "Sold":
            stats["sold"] += int(quantity)
        elif grading_status == "Submitted":
            stats["at_grading"] += int(quantity)
            stats["in_collection"] += int(quantity)
        else:
            stats["in_collection"] += int(quantity)

        if grading_status == "Graded":
            stats["graded"] += int(quantity)

    for row in range(2, 202):
        if ws_sales.cell(row=row, column=1).value is None:
            continue
        sale_price = ws_sales.cell(row=row, column=5).value or 0
        fees = ws_sales.cell(row=row, column=6).value or 0
        shipping = ws_sales.cell(row=row, column=7).value or 0
        purchase = ws_sales.cell(row=row, column=9).value or 0
        stats["total_sale_revenue"] += float(sale_price)
        stats["total_fees"] += float(fees) + float(shipping)
        stats["total_sale_profit"] += float(sale_price) - float(fees) - float(shipping) - float(purchase)

    wb.close()
    return stats


def edit_card(card_id, name, set_name, card_number, language, condition, price, source, quantity):
    wb = load_workbook(EXCEL_PATH)
    ws = wb["Inventar"]
    for row in range(2, 202):
        if ws.cell(row=row, column=1).value == card_id:
            ws.cell(row=row, column=2, value=name)
            ws.cell(row=row, column=3, value=set_name)
            ws.cell(row=row, column=4, value=card_number)
            ws.cell(row=row, column=5, value=language)
            ws.cell(row=row, column=6, value=condition)
            ws.cell(row=row, column=11, value=price)
            ws.cell(row=row, column=11).number_format = '#,##0.00 "EUR"'
            ws.cell(row=row, column=13, value=source)
            ws.cell(row=row, column=14, value=quantity)
            break
    wb.save(EXCEL_PATH)
    wb.close()


def load_card(card_id):
    wb = load_workbook(EXCEL_PATH)
    ws = wb["Inventar"]
    for row in range(2, 202):
        if ws.cell(row=row, column=1).value == card_id:
            card = {
                "id": card_id,
                "name": ws.cell(row=row, column=2).value or "",
                "set": ws.cell(row=row, column=3).value or "",
                "number": ws.cell(row=row, column=4).value or "",
                "language": ws.cell(row=row, column=5).value or "",
                "condition": ws.cell(row=row, column=6).value or "Near Mint",
                "price": ws.cell(row=row, column=11).value or 0,
                "source": ws.cell(row=row, column=13).value or "",
                "quantity": ws.cell(row=row, column=14).value or 1,
            }
            wb.close()
            return card
    wb.close()
    return None


def delete_card(card_id):
    wb = load_workbook(EXCEL_PATH)
    ws = wb["Inventar"]
    for row in range(2, 202):
        if ws.cell(row=row, column=1).value == card_id:
            for col in range(1, 16):
                ws.cell(row=row, column=col).value = None
            break
    ws_sales = wb["Sales"]
    for row in range(2, 202):
        if ws_sales.cell(row=row, column=1).value == card_id:
            for col in range(1, 13):
                ws_sales.cell(row=row, column=col).value = None
            break
    wb.save(EXCEL_PATH)
    wb.close()
    photo = get_photo_path(card_id)
    if photo:
        os.remove(photo)


def _load_config():
    try:
        with open(CONFIG_PATH) as f:
            return json.load(f)
    except (FileNotFoundError, json.JSONDecodeError):
        return {}


def _save_config(config):
    with open(CONFIG_PATH, "w") as f:
        json.dump(config, f)


# ============================================================
# PHOTO OPERATIONS
# ============================================================

PHOTO_DIR = os.path.join(BASE_DIR, "Fotos")
os.makedirs(PHOTO_DIR, exist_ok=True)

URLS_PATH = os.path.join(BASE_DIR, ".card_urls.json")


def _load_urls():
    try:
        with open(URLS_PATH) as f:
            return json.load(f)
    except (FileNotFoundError, json.JSONDecodeError):
        return {}


def _save_url(card_id, url):
    urls = _load_urls()
    urls[str(card_id)] = url
    with open(URLS_PATH, "w") as f:
        json.dump(urls, f)


def get_card_url(card_id):
    return _load_urls().get(str(card_id))


# ============================================================
# EBAY API
# ============================================================

def _load_ebay_config():
    try:
        with open(EBAY_CONFIG_PATH) as f:
            return json.load(f)
    except (FileNotFoundError, json.JSONDecodeError):
        return {}


def _save_ebay_config(cfg):
    with open(EBAY_CONFIG_PATH, "w") as f:
        json.dump(cfg, f)


def _ebay_base(cfg):
    return "https://api.sandbox.ebay.com" if cfg.get("sandbox") else "https://api.ebay.com"


def _ebay_auth_base(cfg):
    return "https://auth.sandbox.ebay.com" if cfg.get("sandbox") else "https://auth.ebay.com"


def _ebay_exchange_code(auth_code):
    cfg = _load_ebay_config()
    creds = base64.b64encode(f"{cfg['client_id']}:{cfg['client_secret']}".encode()).decode()
    data = urllib.parse.urlencode({
        "grant_type": "authorization_code",
        "code": auth_code,
        "redirect_uri": cfg["ru_name"],
    }).encode()
    req = urllib.request.Request(
        f"{_ebay_base(cfg)}/identity/v1/oauth2/token",
        data=data,
        headers={"Authorization": f"Basic {creds}",
                 "Content-Type": "application/x-www-form-urlencoded"},
    )
    with urllib.request.urlopen(req) as resp:
        result = json.loads(resp.read())
    cfg["access_token"] = result["access_token"]
    cfg["refresh_token"] = result.get("refresh_token", "")
    cfg["token_expiry"] = _time.time() + result.get("expires_in", 7200)
    _save_ebay_config(cfg)


def _ebay_refresh():
    cfg = _load_ebay_config()
    creds = base64.b64encode(f"{cfg['client_id']}:{cfg['client_secret']}".encode()).decode()
    data = urllib.parse.urlencode({
        "grant_type": "refresh_token",
        "refresh_token": cfg["refresh_token"],
        "scope": EBAY_SCOPES,
    }).encode()
    req = urllib.request.Request(
        f"{_ebay_base(cfg)}/identity/v1/oauth2/token",
        data=data,
        headers={"Authorization": f"Basic {creds}",
                 "Content-Type": "application/x-www-form-urlencoded"},
    )
    with urllib.request.urlopen(req) as resp:
        result = json.loads(resp.read())
    cfg["access_token"] = result["access_token"]
    cfg["token_expiry"] = _time.time() + result.get("expires_in", 7200)
    _save_ebay_config(cfg)


def _ebay_token():
    cfg = _load_ebay_config()
    if not cfg.get("access_token"):
        return None
    if _time.time() > cfg.get("token_expiry", 0) - 60:
        try:
            _ebay_refresh()
            cfg = _load_ebay_config()
        except Exception:
            return None
    return cfg["access_token"]


def _ebay_api(method, path, body=None):
    cfg = _load_ebay_config()
    token = _ebay_token()
    if not token:
        raise Exception("Not connected to eBay. Go to the Settings tab first.")
    url = f"{_ebay_base(cfg)}{path}"
    payload = json.dumps(body).encode() if body else None
    req = urllib.request.Request(url, data=payload, method=method, headers={
        "Authorization": f"Bearer {token}",
        "Content-Type": "application/json",
        "Accept": "application/json",
        "Content-Language": "de-DE",
    })
    try:
        with urllib.request.urlopen(req) as resp:
            raw = resp.read()
            return json.loads(raw) if raw else {}
    except urllib.error.HTTPError as e:
        err = e.read().decode()
        raise Exception(f"eBay API error ({e.code}): {err}")


def _ebay_upload_image(image_path):
    token = _ebay_token()
    if not token:
        raise Exception("Not connected to eBay.")
    cfg = _load_ebay_config()
    url = "https://api.ebay.com/ws/api.dll"
    if cfg.get("sandbox"):
        url = "https://api.sandbox.ebay.com/ws/api.dll"

    boundary = "----FormBoundary7MA4YWxkTrZu0gW"
    xml_part = (
        '<?xml version="1.0" encoding="utf-8"?>'
        '<UploadSiteHostedPicturesRequest xmlns="urn:ebay:apis:eBLBaseComponents">'
        f'<RequesterCredentials><eBayAuthToken>{token}</eBayAuthToken></RequesterCredentials>'
        '<PictureName>card</PictureName>'
        '</UploadSiteHostedPicturesRequest>'
    )

    with open(image_path, "rb") as f:
        image_data = f.read()

    ext = os.path.splitext(image_path)[1].lower()
    mime = "image/jpeg" if ext in (".jpg", ".jpeg") else "image/png"

    body = (
        f"--{boundary}\r\n"
        f'Content-Disposition: form-data; name="XML Payload"\r\n'
        f"Content-Type: text/xml\r\n\r\n"
        f"{xml_part}\r\n"
        f"--{boundary}\r\n"
        f'Content-Disposition: form-data; name="file"; filename="card{ext}"\r\n'
        f"Content-Type: {mime}\r\n\r\n"
    ).encode() + image_data + f"\r\n--{boundary}--\r\n".encode()

    req = urllib.request.Request(url, data=body, method="POST", headers={
        "X-EBAY-API-SITEID": "77",
        "X-EBAY-API-CALL-NAME": "UploadSiteHostedPictures",
        "X-EBAY-API-COMPATIBILITY-LEVEL": "967",
        "Content-Type": f"multipart/form-data; boundary={boundary}",
    })
    with urllib.request.urlopen(req) as resp:
        raw = resp.read().decode()
    match = re.search(r"<FullURL>(.+?)</FullURL>", raw)
    if match:
        return match.group(1)
    raise Exception(f"Image upload failed: {raw[:300]}")


def _ebay_fetch_policies():
    cfg = _load_ebay_config()
    mp = cfg.get("marketplace", "EBAY_DE")
    for kind in ("fulfillment", "payment", "return"):
        try:
            result = _ebay_api("GET", f"/sell/account/v1/{kind}_policy?marketplace_id={mp}")
            policies = result.get(f"{kind}Policies", [])
            if policies:
                cfg[f"{kind}_policy_id"] = policies[0][f"{kind}PolicyId"]
        except Exception:
            pass
    _save_ebay_config(cfg)


def _ebay_update_shipping(de_cost, eu_cost, intl_cost, handling_days=2):
    cfg = _load_ebay_config()
    policy_id = cfg.get("fulfillment_policy_id")
    if not policy_id:
        raise Exception("No fulfillment policy found. Connect to eBay first.")
    _ebay_api("PUT", f"/sell/account/v1/fulfillment_policy/{policy_id}", {
        "name": "Standard Versand",
        "marketplaceId": cfg.get("marketplace", "EBAY_DE"),
        "handlingTime": {"value": handling_days, "unit": "DAY"},
        "globalShipping": False,
        "shippingOptions": [
            {
                "optionType": "DOMESTIC",
                "costType": "FLAT_RATE",
                "shippingServices": [{
                    "sortOrder": 1,
                    "shippingCarrierCode": "DHL",
                    "shippingServiceCode": "DE_DHLPaket",
                    "shippingCost": {"value": f"{de_cost:.2f}", "currency": "EUR"},
                    "freeShipping": False,
                }],
            },
            {
                "optionType": "INTERNATIONAL",
                "costType": "FLAT_RATE",
                "shipToLocations": {"regionIncluded": [{"regionName": "WORLDWIDE"}]},
                "shippingServices": [
                    {
                        "sortOrder": 1,
                        "shippingCarrierCode": "DHL",
                        "shippingServiceCode": "DE_DHLPaketInternational",
                        "shippingCost": {"value": f"{eu_cost:.2f}", "currency": "EUR"},
                        "freeShipping": False,
                        "shipToLocations": {"regionIncluded": [{"regionName": "EUROPE"}]},
                    },
                    {
                        "sortOrder": 2,
                        "shippingCarrierCode": "DHL",
                        "shippingServiceCode": "DE_DHLPaketInternational",
                        "shippingCost": {"value": f"{intl_cost:.2f}", "currency": "EUR"},
                        "freeShipping": False,
                        "shipToLocations": {"regionIncluded": [{"regionName": "WORLDWIDE"}]},
                    },
                ],
            },
        ],
    })
    cfg["shipping_de"] = de_cost
    cfg["shipping_eu"] = eu_cost
    cfg["shipping_intl"] = intl_cost
    cfg["handling_days"] = handling_days
    _save_ebay_config(cfg)


def ebay_create_draft(card_id, price, quantity):
    card = load_card(card_id)
    if not card:
        raise Exception(f"Card #{card_id} not found")

    wb = load_workbook(EXCEL_PATH)
    ws = wb["Inventar"]
    condition_raw = "Near Mint"
    grading_service = ""
    grade = ""
    for row in range(2, 202):
        if ws.cell(row=row, column=1).value == card_id:
            condition_raw = ws.cell(row=row, column=6).value or "Near Mint"
            grading_service = ws.cell(row=row, column=8).value or ""
            grade = ws.cell(row=row, column=9).value or ""
            break
    wb.close()

    title_parts = [card["name"]]
    if card["set"]:
        title_parts.append(card["set"])
    if card["number"]:
        title_parts.append(card["number"])
    if grading_service and grade:
        title_parts.append(f"{grading_service} {grade}")
    title_parts.append("Pokemon TCG")
    title = " ".join(title_parts)[:80]

    desc_lines = [f"<h2>{card['name']}</h2>",
                  f"<p><b>Set:</b> {card['set']}</p>"]
    if card["number"]:
        desc_lines.append(f"<p><b>Card Number:</b> {card['number']}</p>")
    desc_lines.append(f"<p><b>Language:</b> {card['language']}</p>")
    desc_lines.append(f"<p><b>Condition:</b> {condition_raw}</p>")
    if grading_service:
        desc_lines.append(f"<p><b>Graded:</b> {grading_service} {grade}</p>")
    description = "\n".join(desc_lines)

    sku = f"PKM-{card_id:04d}"

    if grading_service:
        svc_id = GRADING_SERVICE_TO_EBAY.get(grading_service.upper(), "2750123")
        grade_str = str(grade).strip()
        grade_id = GRADE_TO_EBAY.get(grade_str, "275020")
        cond_enum = "LIKE_NEW"
        cond_descriptors = [
            {"name": "27501", "values": [svc_id]},
            {"name": "27502", "values": [grade_id]},
        ]
    else:
        card_cond_id = CONDITION_TO_EBAY_CARD.get(condition_raw, "400010")
        cond_enum = "USED_VERY_GOOD"
        cond_descriptors = [
            {"name": "40001", "values": [card_cond_id]},
        ]

    image_urls = []
    photo_path = get_photo_path(card_id)
    if photo_path:
        img_url = _ebay_upload_image(photo_path)
        image_urls.append(img_url)

    inv_item = {
        "product": {
            "title": title,
            "description": description,
            "imageUrls": image_urls,
            "aspects": {
                "Kartenname": [card["name"]],
                "Set/Erweiterung": [card["set"]],
                "Spiel": ["Pokémon TCG"],
                "Sprache": [card["language"]],
            },
        },
        "condition": cond_enum,
        "conditionDescriptors": cond_descriptors,
        "conditionDescription": f"{grading_service} {grade}" if grading_service else condition_raw,
        "availability": {
            "shipToLocationAvailability": {"quantity": quantity}
        },
    }
    _ebay_api("PUT", f"/sell/inventory/v1/inventory_item/{sku}", inv_item)

    loc_key = _ebay_ensure_location()
    cfg = _load_ebay_config()
    mp = cfg.get("marketplace", "EBAY_DE")
    offer = {
        "sku": sku,
        "marketplaceId": mp,
        "format": "FIXED_PRICE",
        "listingDescription": description,
        "availableQuantity": quantity,
        "categoryId": "183454",
        "merchantLocationKey": loc_key,
        "pricingSummary": {
            "price": {"currency": "EUR", "value": f"{price:.2f}"}
        },
        "listingPolicies": {},
    }
    for kind in ("fulfillment", "payment", "return"):
        pid = cfg.get(f"{kind}_policy_id")
        if pid:
            offer["listingPolicies"][f"{kind}PolicyId"] = pid

    existing_offer_id = None
    try:
        offers_result = _ebay_api("GET", f"/sell/inventory/v1/offer?sku={urllib.parse.quote(sku)}")
        for o in offers_result.get("offers", []):
            if o.get("sku") == sku:
                existing_offer_id = o.get("offerId")
                break
    except Exception:
        pass

    if existing_offer_id:
        _ebay_api("PUT", f"/sell/inventory/v1/offer/{existing_offer_id}", offer)
        return existing_offer_id, None
    else:
        result = _ebay_api("POST", "/sell/inventory/v1/offer", offer)
        return result.get("offerId", "created"), None


def _ebay_ensure_location():
    cfg = _load_ebay_config()
    loc_key = "default_location"
    mp = cfg.get("marketplace", "EBAY_DE")
    mp_to_country = {
        "EBAY_DE": "DE", "EBAY_US": "US", "EBAY_GB": "GB",
        "EBAY_FR": "FR", "EBAY_IT": "IT", "EBAY_ES": "ES",
    }
    country = mp_to_country.get(mp, "DE")
    try:
        _ebay_api("GET", f"/sell/inventory/v1/location/{loc_key}")
        return loc_key
    except Exception:
        pass
    body = {
        "location": {
            "address": {
                "postalCode": "46487",
                "country": country,
            }
        },
        "name": "Default Location",
        "merchantLocationStatus": "ENABLED",
        "locationTypes": ["WAREHOUSE"],
    }
    _ebay_api("POST", f"/sell/inventory/v1/location/{loc_key}", body)
    cfg["location_key"] = loc_key
    _save_ebay_config(cfg)
    return loc_key


def ebay_publish_offer(offer_id):
    result = _ebay_api("POST", f"/sell/inventory/v1/offer/{offer_id}/publish")
    return result.get("listingId", "")


def _extract_auth_code(text):
    text = text.strip()
    if "code=" in text:
        qs = text.split("?", 1)[1] if "?" in text else text
        params = parse_qs(qs)
        code = params.get("code", [None])[0]
        if code:
            return code
    return text


def get_photo_path(card_id):
    for ext in ("jpg", "jpeg", "png", "heic"):
        path = os.path.join(PHOTO_DIR, f"{card_id}.{ext}")
        if os.path.exists(path):
            return path
    return None


def save_photo(card_id, source_path):
    ext = os.path.splitext(source_path)[1].lower()
    if ext in (".heic", ".heif"):
        img = Image.open(source_path)
        dest = os.path.join(PHOTO_DIR, f"{card_id}.jpg")
        img.save(dest, "JPEG", quality=90)
    else:
        dest = os.path.join(PHOTO_DIR, f"{card_id}{ext}")
        shutil.copy2(source_path, dest)
    return dest


def save_photo_bytes(card_id, data, content_type="image/jpeg"):
    ext = ".jpg"
    if "png" in content_type:
        ext = ".png"
    dest = os.path.join(PHOTO_DIR, f"{card_id}{ext}")
    with open(dest, "wb") as f:
        f.write(data)
    return dest


def _get_local_ip():
    try:
        s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
        s.connect(("8.8.8.8", 80))
        ip = s.getsockname()[0]
        s.close()
        return ip
    except Exception:
        return "127.0.0.1"


# ============================================================
# PHOTO UPLOAD WEB SERVER
# ============================================================

UPLOAD_PORT = 8080

_HTML_TEMPLATE = """<!DOCTYPE html>
<html>
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>Pokemon Photo Upload</title>
<style>
  * { box-sizing: border-box; margin: 0; padding: 0; }
  body { font-family: -apple-system, system-ui, sans-serif; background: #1a1a2e; color: #eee; padding: 16px; }
  h1 { font-size: 20px; text-align: center; margin-bottom: 16px; color: #ffcb05; }
  .subtitle { text-align: center; color: #888; font-size: 13px; margin-bottom: 20px; }
  .card { background: #16213e; border-radius: 12px; padding: 14px; margin-bottom: 10px;
          display: flex; align-items: center; justify-content: space-between; }
  .card-info { flex: 1; }
  .card-name { font-weight: 600; font-size: 15px; }
  .card-set { color: #888; font-size: 13px; margin-top: 2px; }
  .card-id { color: #555; font-size: 12px; }
  .has-photo { border-left: 3px solid #4caf50; }
  .no-photo { border-left: 3px solid #f44336; }
  .photo-btn { background: #ffcb05; color: #1a1a2e; border: none; border-radius: 8px;
               padding: 10px 16px; font-weight: 600; font-size: 14px; cursor: pointer; }
  .photo-btn.done { background: #4caf50; color: white; }
  .upload-form { display: none; }
  .preview { max-width: 100%; max-height: 200px; border-radius: 8px; margin-top: 10px; }
  .msg { text-align: center; padding: 20px; color: #4caf50; font-weight: 600; display: none; }
  input[type=file] { display: none; }
</style>
</head>
<body>
<h1>Pokemon Photo Upload</h1>
<p class="subtitle">Tap a card to take a photo</p>
<div id="cards">{{CARDS}}</div>
<script>
function upload(cardId, btn) {
  const input = document.createElement('input');
  input.type = 'file';
  input.accept = 'image/*';
  input.capture = 'environment';
  input.onchange = async () => {
    const file = input.files[0];
    if (!file) return;
    btn.textContent = '...';
    const form = new FormData();
    form.append('photo', file);
    try {
      const resp = await fetch('/upload/' + cardId, { method: 'POST', body: form });
      if (resp.ok) {
        btn.textContent = 'Done';
        btn.classList.add('done');
        btn.closest('.card').classList.remove('no-photo');
        btn.closest('.card').classList.add('has-photo');
      } else {
        btn.textContent = 'Error';
      }
    } catch(e) {
      btn.textContent = 'Error';
    }
  };
  input.click();
}
</script>
</body>
</html>"""


class PhotoUploadHandler(BaseHTTPRequestHandler):
    def log_message(self, format, *args):
        pass

    def do_GET(self):
        if self.path == "/" or self.path == "":
            cards = load_all_cards()
            cards_html = ""
            for c in cards:
                if c["status"] == "Sold":
                    continue
                has = get_photo_path(c["id"]) is not None
                cls = "has-photo" if has else "no-photo"
                btn_cls = "photo-btn done" if has else "photo-btn"
                btn_text = "Replace" if has else "Photo"
                cards_html += (
                    f'<div class="card {cls}">'
                    f'<div class="card-info">'
                    f'<div class="card-name">{c["name"]}</div>'
                    f'<div class="card-set">{c["set"]}</div>'
                    f'<div class="card-id">#{c["id"]:03d}</div>'
                    f'</div>'
                    f'<button class="{btn_cls}" onclick="upload({c["id"]}, this)">{btn_text}</button>'
                    f'</div>'
                )
            html = _HTML_TEMPLATE.replace("{{CARDS}}", cards_html)
            self.send_response(200)
            self.send_header("Content-Type", "text/html; charset=utf-8")
            self.end_headers()
            self.wfile.write(html.encode())
        else:
            self.send_response(404)
            self.end_headers()

    def do_POST(self):
        if self.path.startswith("/upload/"):
            try:
                card_id = int(self.path.split("/")[-1])
            except ValueError:
                self.send_response(400)
                self.end_headers()
                return

            content_length = int(self.headers.get("Content-Length", 0))
            content_type = self.headers.get("Content-Type", "")

            if "multipart/form-data" in content_type:
                boundary = content_type.split("boundary=")[-1].encode()
                body = self.rfile.read(content_length)
                parts = body.split(b"--" + boundary)
                for part in parts:
                    if b"filename=" in part:
                        header_end = part.find(b"\r\n\r\n")
                        if header_end == -1:
                            continue
                        file_data = part[header_end + 4:]
                        if file_data.endswith(b"\r\n"):
                            file_data = file_data[:-2]
                        ct = "image/jpeg"
                        if b"image/png" in part[:header_end]:
                            ct = "image/png"
                        save_photo_bytes(card_id, file_data, ct)
                        self.send_response(200)
                        self.send_header("Content-Type", "text/plain")
                        self.end_headers()
                        self.wfile.write(b"OK")
                        return

            self.send_response(400)
            self.end_headers()
        else:
            self.send_response(404)
            self.end_headers()


class _ReusableHTTPServer(HTTPServer):
    allow_reuse_address = True

    def server_bind(self):
        self.socket.setsockopt(socket.SOL_SOCKET, socket.SO_REUSEADDR, 1)
        self.socket.bind(self.server_address)
        self.server_address = self.socket.getsockname()
        self.server_name = "localhost"
        self.server_port = self.server_address[1]


def start_photo_server():
    server = _ReusableHTTPServer(("0.0.0.0", UPLOAD_PORT), PhotoUploadHandler)
    server.daemon_threads = True
    thread = threading.Thread(target=server.serve_forever, daemon=True)
    thread.start()
    return server


# ============================================================
# AUTO-UPDATE
# ============================================================

SCRIPT_PATH = os.path.join(BASE_DIR, "pokemon_card_manager.py")


def check_for_update():
    req = urllib.request.Request(GITHUB_RAW_URL, headers={"User-Agent": "PokemonCardManager"})
    with urllib.request.urlopen(req, timeout=15) as resp:
        remote_content = resp.read()
    with open(SCRIPT_PATH, "rb") as f:
        local_hash = hashlib.sha256(f.read()).hexdigest()
    remote_hash = hashlib.sha256(remote_content).hexdigest()
    if local_hash == remote_hash:
        return False, VERSION, None
    remote_text = remote_content.decode("utf-8", errors="replace")
    match = re.search(r'^VERSION\s*=\s*"([^"]+)"', remote_text, re.MULTILINE)
    remote_version = match.group(1) if match else "unknown"
    return True, remote_version, remote_content


def apply_update(content):
    with open(SCRIPT_PATH, "wb") as f:
        f.write(content)


def restart_app():
    os.execv(sys.executable, [sys.executable, SCRIPT_PATH])


# ============================================================
# GUI
# ============================================================

class App:
    def __init__(self, root):
        self.root = root
        root.title("Pokemon Card Manager")
        root.resizable(True, True)

        config = _load_config()
        root.geometry(config.get("geometry", "750x600"))
        root.protocol("WM_DELETE_WINDOW", self._on_close)
        root.createcommand("::tk::mac::Quit", self._on_close)
        root.bind("<Destroy>", self._on_destroy)

        notebook = ttk.Notebook(root)
        notebook.pack(fill="both", expand=True, padx=10, pady=10)
        notebook.bind("<MouseWheel>", lambda e: "break")
        notebook.bind("<Button-4>", lambda e: "break")
        notebook.bind("<Button-5>", lambda e: "break")

        # --- Tab 1: Add ---
        add_outer = ttk.Frame(notebook)
        notebook.add(add_outer, text="  Add  ")
        add_outer.columnconfigure(0, weight=1)
        add_outer.rowconfigure(0, weight=1)

        add_frame = ttk.Frame(add_outer, padding=15)
        add_frame.grid(row=0, column=0)

        add_frame.columnconfigure(1, weight=1)

        ttk.Label(add_frame, text="Add Pokemon Card", font=("Helvetica", 16, "bold")).grid(
            row=0, column=0, columnspan=3, pady=(0, 15)
        )

        ttk.Label(add_frame, text="Cardmarket Link:").grid(row=1, column=0, sticky="e", padx=(0, 10), pady=4)
        self.url_entry = ttk.Entry(add_frame, width=45)
        self.url_entry.grid(row=1, column=1, sticky="ew", pady=4)
        ttk.Button(add_frame, text="Fetch", width=8, command=self.on_fetch).grid(row=1, column=2, padx=(5, 0), pady=4)

        self.card_info_label = ttk.Label(add_frame, text="", font=("Helvetica", 12), foreground="blue")
        self.card_info_label.grid(row=2, column=0, columnspan=3, pady=(5, 10))

        ttk.Label(add_frame, text="Language:").grid(row=3, column=0, sticky="e", padx=(0, 10), pady=4)
        self.lang_cb = ttk.Combobox(add_frame, values=LANGUAGES, width=10, state="readonly")
        self.lang_cb.set("EN")
        self.lang_cb.grid(row=3, column=1, sticky="w", pady=4)

        ttk.Label(add_frame, text="Condition:").grid(row=4, column=0, sticky="e", padx=(0, 10), pady=4)
        self.condition_cb = ttk.Combobox(add_frame, values=CONDITIONS, width=15, state="readonly")
        self.condition_cb.set("Near Mint")
        self.condition_cb.grid(row=4, column=1, sticky="w", pady=4)

        ttk.Label(add_frame, text="Quantity:").grid(row=5, column=0, sticky="e", padx=(0, 10), pady=4)
        self.qty_entry = ttk.Entry(add_frame, width=8)
        self.qty_entry.insert(0, "1")
        self.qty_entry.grid(row=5, column=1, sticky="w", pady=4)

        ttk.Label(add_frame, text="Price per Unit (EUR):").grid(row=6, column=0, sticky="e", padx=(0, 10), pady=4)
        self.price_entry = ttk.Entry(add_frame, width=15)
        self.price_entry.grid(row=6, column=1, sticky="w", pady=4)

        ttk.Label(add_frame, text="Source:").grid(row=7, column=0, sticky="e", padx=(0, 10), pady=4)
        self.source_cb = ttk.Combobox(add_frame, values=SOURCES, width=15, state="readonly")
        self.source_cb.set("Cardmarket")
        self.source_cb.grid(row=7, column=1, sticky="w", pady=4)

        # --- Grading (optional) ---
        sep2 = ttk.Separator(add_frame, orient="horizontal")
        sep2.grid(row=8, column=0, columnspan=3, sticky="ew", pady=10)

        ttk.Label(add_frame, text="Grading:").grid(row=9, column=0, sticky="e", padx=(0, 10), pady=4)
        self.grading_cb = ttk.Combobox(
            add_frame, values=["Not Graded", "Graded"], width=15, state="readonly"
        )
        self.grading_cb.set("Not Graded")
        self.grading_cb.grid(row=9, column=1, sticky="w", pady=4)
        self.grading_cb.bind("<<ComboboxSelected>>", self._toggle_grading_fields)

        self.grading_detail_frame = ttk.Frame(add_frame)
        self.grading_detail_frame.grid(row=10, column=0, columnspan=3, sticky="ew")

        ttk.Label(self.grading_detail_frame, text="Service:").grid(row=0, column=0, sticky="e", padx=(0, 10), pady=4)
        self.grading_service_cb = ttk.Combobox(
            self.grading_detail_frame, values=["PSA", "CGC", "Beckett", "ACE", "TAG", "Other"],
            width=10, state="readonly"
        )
        self.grading_service_cb.set("PSA")
        self.grading_service_cb.grid(row=0, column=1, sticky="w", pady=4)

        ttk.Label(self.grading_detail_frame, text="Grade:").grid(row=0, column=2, sticky="e", padx=(15, 10), pady=4)
        self.grade_entry = ttk.Entry(self.grading_detail_frame, width=6)
        self.grade_entry.grid(row=0, column=3, sticky="w", pady=4)

        ttk.Label(self.grading_detail_frame, text="Cert #:").grid(row=0, column=4, sticky="e", padx=(15, 10), pady=4)
        self.cert_entry = ttk.Entry(self.grading_detail_frame, width=12)
        self.cert_entry.grid(row=0, column=5, sticky="w", pady=4)

        self.grading_detail_frame.grid_remove()

        ttk.Button(add_frame, text="Save", command=self.on_save).grid(
            row=11, column=0, columnspan=3, pady=(15, 5)
        )

        self.add_status = ttk.Label(add_frame, text="Paste a Cardmarket link and click Fetch", foreground="gray")
        self.add_status.grid(row=12, column=0, columnspan=3, pady=(5, 0))

        self.fetched_data = None
        self.url_entry.focus()
        root.bind("<Return>", lambda e: self._handle_enter())

        # --- Tab 2: My Collection ---
        col_frame = ttk.Frame(notebook, padding=15)
        notebook.add(col_frame, text="  My Collection  ")

        # Search bar
        search_frame = ttk.Frame(col_frame)
        search_frame.pack(fill="x", pady=(0, 10))
        ttk.Label(search_frame, text="Search:").pack(side="left", padx=(0, 5))
        self.search_var = tk.StringVar()
        self.search_var.trace_add("write", lambda *_: self.refresh_collection())
        search_entry = ttk.Entry(search_frame, textvariable=self.search_var, width=30)
        search_entry.pack(side="left")
        ttk.Label(search_frame, text="  Filter:", foreground="gray").pack(side="left", padx=(15, 5))

        saved_filter = config.get("filter", "All")
        self.filter_var = tk.StringVar(value=saved_filter)
        for text in ["All", "Collection", "At Grading", "Sold"]:
            ttk.Radiobutton(search_frame, text=text, variable=self.filter_var, value=text,
                            command=self.refresh_collection).pack(side="left", padx=3)

        local_ip = _get_local_ip()
        phone_url = f"http://{local_ip}:{UPLOAD_PORT}"
        url_frame = ttk.Frame(col_frame)
        url_frame.pack(side="bottom", fill="x", pady=(8, 0))
        ttk.Label(url_frame, text="Phone upload:", foreground="gray").pack(side="left")
        url_label = ttk.Label(url_frame, text=phone_url, foreground="#ffcb05", cursor="hand2")
        url_label.pack(side="left", padx=(5, 0))
        url_label.bind("<Button-1>", lambda e: self.root.clipboard_clear() or self.root.clipboard_append(phone_url))

        # Treeview
        tree_frame = ttk.Frame(col_frame)
        tree_frame.pack(side="left", fill="both", expand=True)

        columns = ("id", "photo", "name", "set", "qty", "price", "status", "grading")
        self.tree = ttk.Treeview(tree_frame, columns=columns, show="headings", height=18)

        self.tree.heading("id", text="ID")
        self.tree.heading("photo", text="\U0001f4f7")
        self.tree.heading("name", text="Card Name")
        self.tree.heading("set", text="Set")
        self.tree.heading("qty", text="Qty")
        self.tree.heading("price", text="Price/Unit")
        self.tree.heading("status", text="Status")
        self.tree.heading("grading", text="Grading")

        default_widths = {"id": 40, "photo": 35, "name": 160, "set": 150, "qty": 40,
                          "price": 80, "status": 100, "grading": 120}
        saved_widths = config.get("column_widths", {})
        anchors = {"id": "center", "photo": "center", "qty": "center", "price": "e", "status": "center", "grading": "center"}
        for col, w in default_widths.items():
            self.tree.column(col, width=saved_widths.get(col, w), anchor=anchors.get(col, "w"))

        scrollbar = ttk.Scrollbar(tree_frame, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=scrollbar.set)

        self.tree.pack(side="left", fill="both", expand=True)
        self.tree.bind("<Double-1>", self._on_tree_double_click)
        scrollbar.pack(side="left", fill="y")

        btn_frame = ttk.Frame(col_frame)
        btn_frame.pack(side="right", fill="y", padx=(10, 0))

        ttk.Button(btn_frame, text="Delete Photo", width=18, command=self.on_delete_photo).pack(pady=5)
        ttk.Separator(btn_frame, orient="horizontal").pack(fill="x", pady=8)
        ttk.Button(btn_frame, text="Edit Card", width=18, command=self.on_edit).pack(pady=5)
        ttk.Button(btn_frame, text="Mark Sold", width=18, command=self.on_mark_sold).pack(pady=5)
        ttk.Button(btn_frame, text="List on eBay", width=18, command=self.on_ebay_list).pack(pady=5)
        ttk.Separator(btn_frame, orient="horizontal").pack(fill="x", pady=8)
        ttk.Button(btn_frame, text="Send to Grading", width=18, command=self.on_send_grading).pack(pady=5)
        ttk.Button(btn_frame, text="Grading Returned", width=18, command=self.on_grading_returned).pack(pady=5)
        ttk.Button(btn_frame, text="Delete Card", width=18, command=self.on_delete).pack(pady=20)
        ttk.Button(btn_frame, text="Refresh", width=18, command=self.refresh_collection).pack(pady=5)

        # --- Tab 3: Statistics ---
        stats_frame = ttk.Frame(notebook, padding=20)
        notebook.add(stats_frame, text="  Statistics  ")

        ttk.Label(stats_frame, text="Portfolio Statistics", font=("Helvetica", 16, "bold")).pack(pady=(0, 20))

        self.stats_container = ttk.Frame(stats_frame)
        self.stats_container.pack(fill="both", expand=True)

        # --- Tab 4: Settings ---
        settings_frame = ttk.Frame(notebook, padding=20)
        notebook.add(settings_frame, text="  Settings  ")

        settings_canvas = tk.Canvas(settings_frame, highlightthickness=0)
        settings_scroll = ttk.Scrollbar(settings_frame, orient="vertical", command=settings_canvas.yview)
        settings_inner = ttk.Frame(settings_canvas)
        settings_inner.bind("<Configure>", lambda e: settings_canvas.configure(scrollregion=settings_canvas.bbox("all")))
        settings_canvas.create_window((0, 0), window=settings_inner, anchor="nw")
        settings_canvas.configure(yscrollcommand=settings_scroll.set)
        settings_canvas.pack(side="left", fill="both", expand=True)
        settings_scroll.pack(side="right", fill="y")

        # -- eBay Connection --
        ebay_frame = ttk.LabelFrame(settings_inner, text="eBay Connection", padding=15)
        ebay_frame.pack(fill="x", pady=(0, 15), padx=5)

        cfg = _load_ebay_config()

        ttk.Label(ebay_frame, text="Client ID:").grid(row=0, column=0, sticky="e", padx=(0, 10), pady=4)
        self.cfg_client_id = ttk.Entry(ebay_frame, width=45)
        self.cfg_client_id.insert(0, cfg.get("client_id", ""))
        self.cfg_client_id.grid(row=0, column=1, pady=4, sticky="w")

        ttk.Label(ebay_frame, text="Client Secret:").grid(row=1, column=0, sticky="e", padx=(0, 10), pady=4)
        self.cfg_client_secret = ttk.Entry(ebay_frame, width=45, show="*")
        self.cfg_client_secret.insert(0, cfg.get("client_secret", ""))
        self.cfg_client_secret.grid(row=1, column=1, pady=4, sticky="w")

        ttk.Label(ebay_frame, text="RuName:").grid(row=2, column=0, sticky="e", padx=(0, 10), pady=4)
        self.cfg_ru_name = ttk.Entry(ebay_frame, width=45)
        self.cfg_ru_name.insert(0, cfg.get("ru_name", ""))
        self.cfg_ru_name.grid(row=2, column=1, pady=4, sticky="w")

        ttk.Label(ebay_frame, text="Marketplace:").grid(row=3, column=0, sticky="e", padx=(0, 10), pady=4)
        self.cfg_marketplace = ttk.Combobox(ebay_frame, values=["EBAY_DE", "EBAY_US", "EBAY_GB", "EBAY_FR", "EBAY_IT", "EBAY_ES"],
                                            width=12, state="readonly")
        self.cfg_marketplace.set(cfg.get("marketplace", "EBAY_DE"))
        self.cfg_marketplace.grid(row=3, column=1, sticky="w", pady=4)

        self.ebay_status_label = ttk.Label(ebay_frame, text="")
        self.ebay_status_label.grid(row=5, column=0, columnspan=2, pady=(5, 0))
        if cfg.get("access_token"):
            self.ebay_status_label.config(text="Status: Connected", foreground="green")
        else:
            self.ebay_status_label.config(text="Status: Not connected", foreground="red")

        ebay_btn_row = ttk.Frame(ebay_frame)
        ebay_btn_row.grid(row=4, column=0, columnspan=2, pady=(10, 0))
        ttk.Button(ebay_btn_row, text="Save", command=self._settings_ebay_save).pack(side="left", padx=5)
        ttk.Button(ebay_btn_row, text="Save & Connect", command=self._settings_ebay_connect).pack(side="left", padx=5)

        # -- Shipping --
        ship_frame = ttk.LabelFrame(settings_inner, text="Shipping / Versand", padding=15)
        ship_frame.pack(fill="x", pady=(0, 15), padx=5)

        ttk.Label(ship_frame, text="Deutschland (EUR):").grid(row=0, column=0, sticky="e", padx=(0, 10), pady=4)
        self.cfg_ship_de = ttk.Entry(ship_frame, width=10)
        self.cfg_ship_de.insert(0, f"{cfg.get('shipping_de', 6.19):.2f}")
        self.cfg_ship_de.grid(row=0, column=1, sticky="w", pady=4)

        ttk.Label(ship_frame, text="EU (EUR):").grid(row=1, column=0, sticky="e", padx=(0, 10), pady=4)
        self.cfg_ship_eu = ttk.Entry(ship_frame, width=10)
        self.cfg_ship_eu.insert(0, f"{cfg.get('shipping_eu', 14.49):.2f}")
        self.cfg_ship_eu.grid(row=1, column=1, sticky="w", pady=4)

        ttk.Label(ship_frame, text="International (EUR):").grid(row=2, column=0, sticky="e", padx=(0, 10), pady=4)
        self.cfg_ship_intl = ttk.Entry(ship_frame, width=10)
        self.cfg_ship_intl.insert(0, f"{cfg.get('shipping_intl', 26.49):.2f}")
        self.cfg_ship_intl.grid(row=2, column=1, sticky="w", pady=4)

        ttk.Label(ship_frame, text="Handling Time (Days):").grid(row=3, column=0, sticky="e", padx=(0, 10), pady=4)
        self.cfg_handling = ttk.Combobox(ship_frame, values=["1", "2", "3", "4", "5"], width=5, state="readonly")
        self.cfg_handling.set(str(cfg.get("handling_days", 2)))
        self.cfg_handling.grid(row=3, column=1, sticky="w", pady=4)

        self.ship_status = ttk.Label(ship_frame, text="")
        self.ship_status.grid(row=5, column=0, columnspan=2, pady=(5, 0))

        ttk.Button(ship_frame, text="Save Shipping", command=self._settings_save_shipping).grid(
            row=4, column=0, columnspan=2, pady=(10, 0))

        # -- App --
        app_frame = ttk.LabelFrame(settings_inner, text="App", padding=15)
        app_frame.pack(fill="x", pady=(0, 15), padx=5)

        ttk.Label(app_frame, text=f"Version: v{VERSION}").pack(anchor="w")
        ttk.Button(app_frame, text="Check for Updates", command=self.on_check_update).pack(anchor="w", pady=(10, 0))

        notebook.bind("<<NotebookTabChanged>>", lambda e: self._on_tab_change(notebook))
        self.refresh_collection()

    def _save_state(self):
        config = _load_config()
        config["geometry"] = self.root.geometry()
        config["filter"] = self.filter_var.get()
        config["column_widths"] = {
            col: self.tree.column(col, "width")
            for col in ("id", "photo", "name", "set", "qty", "price", "status", "grading")
        }
        _save_config(config)

    def _on_close(self):
        self._save_state()
        self.root.destroy()

    def _on_destroy(self, event):
        if event.widget is self.root:
            self._save_state()

    def _on_tab_change(self, notebook):
        tab = notebook.tab(notebook.select(), "text").strip()
        if tab == "My Collection":
            self.refresh_collection()
        elif tab == "Statistics":
            self.refresh_statistics()

    def refresh_statistics(self):
        for widget in self.stats_container.winfo_children():
            widget.destroy()

        try:
            s = load_statistics()
        except Exception:
            ttk.Label(self.stats_container, text="Could not load statistics.").pack()
            return

        left = ttk.LabelFrame(self.stats_container, text="Collection", padding=15)
        left.grid(row=0, column=0, sticky="nsew", padx=(0, 10), pady=5)

        right = ttk.LabelFrame(self.stats_container, text="Sales & Profit", padding=15)
        right.grid(row=0, column=1, sticky="nsew", padx=(10, 0), pady=5)

        self.stats_container.columnconfigure(0, weight=1)
        self.stats_container.columnconfigure(1, weight=1)

        collection_rows = [
            ("Total Items", str(s["total_items"])),
            ("In Collection", str(s["in_collection"])),
            ("At Grading", str(s["at_grading"])),
            ("Graded", str(s["graded"])),
            ("Sold", str(s["sold"])),
            ("", ""),
            ("Total Spent", f"{s['total_spent']:.2f} EUR"),
        ]

        for i, (label, value) in enumerate(collection_rows):
            if not label:
                ttk.Separator(left, orient="horizontal").grid(row=i, column=0, columnspan=2, sticky="ew", pady=8)
                continue
            ttk.Label(left, text=label, font=("Helvetica", 11)).grid(row=i, column=0, sticky="w", pady=3)
            ttk.Label(left, text=value, font=("Helvetica", 11, "bold")).grid(
                row=i, column=1, sticky="e", padx=(20, 0), pady=3
            )

        sales_rows = [
            ("Total Sales", str(s["sold"])),
            ("Gross Revenue", f"{s['total_sale_revenue']:.2f} EUR"),
            ("Total Fees & Shipping", f"{s['total_fees']:.2f} EUR"),
            ("", ""),
            ("Net Profit (Sales)", f"{s['total_sale_profit']:.2f} EUR"),
        ]

        for i, (label, value) in enumerate(sales_rows):
            if not label:
                ttk.Separator(right, orient="horizontal").grid(row=i, column=0, columnspan=2, sticky="ew", pady=8)
                continue
            ttk.Label(right, text=label, font=("Helvetica", 11)).grid(row=i, column=0, sticky="w", pady=3)
            kwargs = {}
            if "Profit" in label:
                val_str = value.replace(" EUR", "")
                try:
                    kwargs["foreground"] = "#4caf50" if float(val_str) >= 0 else "#f44336"
                except ValueError:
                    pass
            ttk.Label(right, text=value, font=("Helvetica", 11, "bold"), **kwargs).grid(
                row=i, column=1, sticky="e", padx=(20, 0), pady=3
            )

    def _handle_enter(self):
        if self.url_entry.get().strip() and not self.fetched_data:
            self.on_fetch()
        else:
            self.on_save()

    def _toggle_grading_fields(self, event=None):
        if self.grading_cb.get() == "Graded":
            self.grading_detail_frame.grid()
        else:
            self.grading_detail_frame.grid_remove()

    def on_fetch(self):
        url = self.url_entry.get().strip()
        if not url:
            messagebox.showwarning("Missing Field", "Please paste a Cardmarket link.")
            return

        result = parse_cardmarket_url(url)
        if not result:
            messagebox.showwarning("Invalid Link", "Could not parse this URL.\nMake sure it's a Cardmarket link.")
            return

        self.fetched_data = result
        display = f"{result['name']}  —  {result['set']}"
        self.card_info_label.config(text=display)
        self.add_status.config(text="Card found! Enter price and save.", foreground="blue")
        self.price_entry.focus()

    def on_save(self):
        if not self.fetched_data:
            messagebox.showwarning("No Card", "Please fetch a card from a Cardmarket link first.")
            return

        d = self.fetched_data
        language = self.lang_cb.get()
        condition = self.condition_cb.get()
        price_str = self.price_entry.get().strip().replace(",", ".")
        source = self.source_cb.get()
        qty_str = self.qty_entry.get().strip()

        try:
            price = float(price_str) if price_str else 0.0
        except ValueError:
            messagebox.showwarning("Invalid Input", "Purchase price must be a number.")
            return

        try:
            quantity = int(qty_str) if qty_str else 1
            if quantity < 1:
                raise ValueError
        except ValueError:
            messagebox.showwarning("Invalid Input", "Quantity must be a positive whole number.")
            return

        grading_status = self.grading_cb.get()
        grading_service = ""
        grade = None
        cert_number = ""
        if grading_status == "Graded":
            grading_service = self.grading_service_cb.get()
            grade_str = self.grade_entry.get().strip()
            if grade_str:
                try:
                    grade = parse_grade(grade_str)
                except ValueError:
                    messagebox.showwarning("Invalid Input", "Grade must be a number (e.g. 9.5) or P10/B10")
                    return
            cert_number = self.cert_entry.get().strip()

        try:
            card_id = save_card(
                d["name"], d["set"], d["number"],
                language, condition, price, source,
                quantity, grading_status, grading_service, grade, cert_number,
                url=self.url_entry.get().strip(),
            )
        except Exception as e:
            messagebox.showerror("Error", f"Failed to save:\n{e}")
            return

        total = price * quantity
        self.add_status.config(
            text=f"Saved #{card_id:03d} {d['name']} ({d['set']}) x{quantity} = {total:.2f} EUR",
            foreground="green",
        )
        self.url_entry.delete(0, tk.END)
        self.price_entry.delete(0, tk.END)
        self.qty_entry.delete(0, tk.END)
        self.qty_entry.insert(0, "1")
        self.card_info_label.config(text="")
        self.condition_cb.set("Near Mint")
        self.grading_cb.set("Not Graded")
        self.grading_detail_frame.grid_remove()
        self.grade_entry.delete(0, tk.END)
        self.cert_entry.delete(0, tk.END)
        self.fetched_data = None
        self.url_entry.focus()

    def refresh_collection(self):
        for item in self.tree.get_children():
            self.tree.delete(item)
        try:
            cards = load_all_cards()
        except Exception:
            return

        cards.sort(key=lambda c: c["id"])
        search = self.search_var.get().lower()
        status_filter = self.filter_var.get()

        for c in cards:
            if search:
                searchable = f"{c['name']} {c['set']} {c['language']}".lower()
                if search not in searchable:
                    continue

            if status_filter == "Collection" and (c["status"] != "Collection" or c["grading_status"] == "Submitted"):
                continue
            elif status_filter == "Sold" and c["status"] != "Sold":
                continue
            elif status_filter == "At Grading" and c["grading_status"] != "Submitted":
                continue

            price_str = f"{c['price']:.2f} EUR" if c["price"] is not None else ""
            qty = int(c["quantity"])
            grading = c["grading_status"]
            if c["grading_service"]:
                grading += f" ({c['grading_service']}"
                if c["grade"]:
                    grading += f" {c['grade']}"
                grading += ")"

            if c["status"] == "Sold":
                tag = "sold"
                status_display = "\U0001f534 Sold"
            elif c["grading_status"] == "Submitted":
                tag = "grading"
                status_display = "\U0001f7e0 Grading"
            else:
                tag = "collection"
                status_display = "\U0001f7e2 Collection"

            photo_icon = "\U0001f4f7" if get_photo_path(c["id"]) else ""

            self.tree.insert("", "end", iid=str(c["id"]), values=(
                c["id"], photo_icon, c["name"], c["set"], qty,
                price_str, status_display, grading,
            ), tags=(tag,))

    def _get_selected_id(self):
        sel = self.tree.selection()
        if not sel:
            messagebox.showwarning("No Selection", "Please select a card first.")
            return None
        return int(sel[0])

    def on_edit(self):
        card_id = self._get_selected_id()
        if card_id is None:
            return

        card = load_card(card_id)
        if card is None:
            messagebox.showerror("Error", f"Card #{card_id} not found.")
            return

        win = tk.Toplevel(self.root)
        win.title(f"Edit Card #{card_id}")
        win.resizable(False, False)
        win.grab_set()

        frame = ttk.Frame(win, padding=15)
        frame.pack()

        fields = {}

        ttk.Label(frame, text="Card Name:").grid(row=0, column=0, sticky="e", padx=(0, 10), pady=4)
        fields["name"] = ttk.Entry(frame, width=30)
        fields["name"].insert(0, card["name"])
        fields["name"].grid(row=0, column=1, pady=4)

        ttk.Label(frame, text="Set / Edition:").grid(row=1, column=0, sticky="e", padx=(0, 10), pady=4)
        fields["set"] = ttk.Combobox(frame, values=SETS, width=28)
        fields["set"].set(card["set"])
        fields["set"].grid(row=1, column=1, pady=4)

        ttk.Label(frame, text="Code:").grid(row=2, column=0, sticky="e", padx=(0, 10), pady=4)
        fields["number"] = ttk.Entry(frame, width=15)
        fields["number"].insert(0, card["number"])
        fields["number"].grid(row=2, column=1, sticky="w", pady=4)

        ttk.Label(frame, text="Language:").grid(row=3, column=0, sticky="e", padx=(0, 10), pady=4)
        fields["language"] = ttk.Combobox(frame, values=LANGUAGES, width=10, state="readonly")
        fields["language"].set(card["language"])
        fields["language"].grid(row=3, column=1, sticky="w", pady=4)

        ttk.Label(frame, text="Condition:").grid(row=4, column=0, sticky="e", padx=(0, 10), pady=4)
        fields["condition"] = ttk.Combobox(frame, values=CONDITIONS, width=15, state="readonly")
        fields["condition"].set(card["condition"])
        fields["condition"].grid(row=4, column=1, sticky="w", pady=4)

        ttk.Label(frame, text="Quantity:").grid(row=5, column=0, sticky="e", padx=(0, 10), pady=4)
        fields["quantity"] = ttk.Entry(frame, width=8)
        fields["quantity"].insert(0, str(card["quantity"]))
        fields["quantity"].grid(row=5, column=1, sticky="w", pady=4)

        ttk.Label(frame, text="Price per Unit (EUR):").grid(row=6, column=0, sticky="e", padx=(0, 10), pady=4)
        fields["price"] = ttk.Entry(frame, width=15)
        fields["price"].insert(0, str(card["price"]))
        fields["price"].grid(row=6, column=1, sticky="w", pady=4)

        ttk.Label(frame, text="Source:").grid(row=7, column=0, sticky="e", padx=(0, 10), pady=4)
        fields["source"] = ttk.Combobox(frame, values=SOURCES, width=15, state="readonly")
        fields["source"].set(card["source"] or "Cardmarket")
        fields["source"].grid(row=7, column=1, sticky="w", pady=4)

        def confirm():
            price_str = fields["price"].get().strip().replace(",", ".")
            try:
                price = float(price_str) if price_str else 0.0
            except ValueError:
                messagebox.showwarning("Invalid Input", "Price must be a number.")
                return
            try:
                qty = int(fields["quantity"].get().strip() or "1")
                if qty < 1:
                    raise ValueError
            except ValueError:
                messagebox.showwarning("Invalid Input", "Quantity must be a positive whole number.")
                return
            edit_card(
                card_id,
                fields["name"].get().strip(),
                fields["set"].get(),
                fields["number"].get().strip(),
                fields["language"].get(),
                fields["condition"].get(),
                price,
                fields["source"].get(),
                qty,
            )
            win.destroy()
            self.refresh_collection()

        ttk.Button(frame, text="Save Changes", command=confirm).grid(
            row=8, column=0, columnspan=2, pady=(15, 0)
        )

    def on_mark_sold(self):
        card_id = self._get_selected_id()
        if card_id is None:
            return

        card = load_card(card_id)
        if card is None:
            return
        total_qty = int(card.get("quantity", 1))

        win = tk.Toplevel(self.root)
        win.title("Mark as Sold")
        win.resizable(False, False)
        win.grab_set()

        frame = ttk.Frame(win, padding=15)
        frame.pack()

        ttk.Label(frame, text="Quantity to Sell:").grid(row=0, column=0, sticky="e", padx=(0, 10), pady=4)
        qty_entry = ttk.Entry(frame, width=8)
        qty_entry.insert(0, str(total_qty))
        qty_entry.grid(row=0, column=1, sticky="w", pady=4)
        ttk.Label(frame, text=f"  / {total_qty} available", foreground="gray").grid(row=0, column=1, sticky="e", padx=(80, 0), pady=4)

        ttk.Label(frame, text="Sale Price (EUR):").grid(row=1, column=0, sticky="e", padx=(0, 10), pady=4)
        sale_price_entry = ttk.Entry(frame, width=15)
        sale_price_entry.grid(row=1, column=1, pady=4)

        ttk.Label(frame, text="Platform:").grid(row=2, column=0, sticky="e", padx=(0, 10), pady=4)
        platform_cb = ttk.Combobox(frame, values=["Cardmarket", "eBay"], width=13, state="readonly")
        platform_cb.set("Cardmarket")
        platform_cb.grid(row=2, column=1, pady=4)

        ttk.Label(frame, text="Fees (EUR):").grid(row=3, column=0, sticky="e", padx=(0, 10), pady=4)
        fees_entry = ttk.Entry(frame, width=15)
        fees_entry.insert(0, "0")
        fees_entry.grid(row=3, column=1, pady=4)

        def confirm():
            try:
                sell_qty = int(qty_entry.get().strip())
                sp = float(sale_price_entry.get().strip().replace(",", "."))
                fees = float(fees_entry.get().strip().replace(",", ".") or "0")
            except ValueError:
                messagebox.showwarning("Invalid Input", "Quantity must be a whole number, prices must be numbers.")
                return
            if sell_qty < 1 or sell_qty > total_qty:
                messagebox.showwarning("Invalid Quantity", f"Must be between 1 and {total_qty}.")
                return
            mark_card_sold(card_id, sp, platform_cb.get(), fees, sell_qty)
            win.destroy()
            self.refresh_collection()

        ttk.Button(frame, text="Confirm Sale", command=confirm).grid(
            row=4, column=0, columnspan=2, pady=(10, 0)
        )
        sale_price_entry.focus()

    def on_send_grading(self):
        card_id = self._get_selected_id()
        if card_id is None:
            return

        win = tk.Toplevel(self.root)
        win.title("Send to Grading")
        win.resizable(False, False)
        win.grab_set()

        frame = ttk.Frame(win, padding=15)
        frame.pack()

        ttk.Label(frame, text="Grading Service:").grid(row=0, column=0, sticky="e", padx=(0, 10), pady=4)
        service_cb = ttk.Combobox(frame, values=["PSA", "CGC", "Beckett", "ACE", "TAG", "Other"], width=13, state="readonly")
        service_cb.set("PSA")
        service_cb.grid(row=0, column=1, pady=4)

        def confirm():
            submit_for_grading(card_id, service_cb.get())
            win.destroy()
            self.refresh_collection()

        ttk.Button(frame, text="Confirm", command=confirm).grid(
            row=1, column=0, columnspan=2, pady=(10, 0)
        )

    def on_grading_returned(self):
        card_id = self._get_selected_id()
        if card_id is None:
            return

        win = tk.Toplevel(self.root)
        win.title("Grading Returned")
        win.resizable(False, False)
        win.grab_set()

        frame = ttk.Frame(win, padding=15)
        frame.pack()

        ttk.Label(frame, text="Grade:").grid(row=0, column=0, sticky="e", padx=(0, 10), pady=4)
        grade_entry = ttk.Entry(frame, width=10)
        grade_entry.grid(row=0, column=1, pady=4)

        ttk.Label(frame, text="Cert Number:").grid(row=1, column=0, sticky="e", padx=(0, 10), pady=4)
        cert_entry = ttk.Entry(frame, width=15)
        cert_entry.grid(row=1, column=1, pady=4)

        def confirm():
            grade_str = grade_entry.get().strip()
            try:
                grade = parse_grade(grade_str) if grade_str else 0
            except ValueError:
                messagebox.showwarning("Invalid Input", "Grade must be a number (e.g. 9.5) or P10/B10")
                return
            grading_returned(card_id, grade, cert_entry.get().strip())
            win.destroy()
            self.refresh_collection()

        ttk.Button(frame, text="Confirm", command=confirm).grid(
            row=2, column=0, columnspan=2, pady=(10, 0)
        )
        grade_entry.focus()

    def _settings_ebay_save(self):
        cfg = _load_ebay_config()
        cfg["client_id"] = self.cfg_client_id.get().strip()
        cfg["client_secret"] = self.cfg_client_secret.get().strip()
        cfg["ru_name"] = self.cfg_ru_name.get().strip()
        cfg["marketplace"] = self.cfg_marketplace.get()
        _save_ebay_config(cfg)
        self.ebay_status_label.config(text="Settings saved.", foreground="blue")

    def _settings_ebay_connect(self):
        self._settings_ebay_save()
        cfg = _load_ebay_config()

        if not cfg.get("client_id") or not cfg.get("client_secret") or not cfg.get("ru_name"):
            messagebox.showwarning("Missing Fields", "Please fill in Client ID, Client Secret, and RuName.")
            return

        auth_url = (
            f"{_ebay_auth_base(cfg)}/oauth2/authorize"
            f"?client_id={urllib.parse.quote(cfg['client_id'])}"
            f"&response_type=code"
            f"&redirect_uri={urllib.parse.quote(cfg['ru_name'])}"
            f"&scope={urllib.parse.quote(EBAY_SCOPES)}"
        )
        webbrowser.open(auth_url)
        self.ebay_status_label.config(text="Waiting for authorization...", foreground="orange")

        paste_url = simpledialog.askstring(
            "eBay Authorization",
            "1. Log in to eBay in the browser window that just opened.\n"
            "2. After granting access, you will be redirected.\n"
            "3. Copy the FULL URL from your browser's address bar\n"
            "   and paste it here:\n",
            parent=self.root,
        )
        if not paste_url:
            self.ebay_status_label.config(text="Status: Not connected", foreground="red")
            return

        try:
            auth_code = _extract_auth_code(paste_url)
            _ebay_exchange_code(auth_code)
            _ebay_fetch_policies()
            self.ebay_status_label.config(text="Status: Connected!", foreground="green")
            messagebox.showinfo("eBay", "Successfully connected to eBay!")
        except Exception as e:
            self.ebay_status_label.config(text="Connection failed", foreground="red")
            messagebox.showerror("eBay Error", str(e))

    def _settings_save_shipping(self):
        try:
            de = float(self.cfg_ship_de.get().replace(",", "."))
            eu = float(self.cfg_ship_eu.get().replace(",", "."))
            intl = float(self.cfg_ship_intl.get().replace(",", "."))
            handling = int(self.cfg_handling.get())
        except ValueError:
            messagebox.showwarning("Invalid Input", "Please enter valid numbers for all shipping costs.")
            return

        cfg = _load_ebay_config()
        if not cfg.get("access_token"):
            cfg["shipping_de"] = de
            cfg["shipping_eu"] = eu
            cfg["shipping_intl"] = intl
            cfg["handling_days"] = handling
            _save_ebay_config(cfg)
            self.ship_status.config(text="Saved locally (not connected to eBay).", foreground="blue")
            return

        self.ship_status.config(text="Updating eBay policy...", foreground="orange")
        self.root.update_idletasks()

        try:
            _ebay_update_shipping(de, eu, intl, handling)
            self.ship_status.config(text="Shipping costs updated!", foreground="green")
        except Exception as e:
            self.ship_status.config(text="Update failed.", foreground="red")
            messagebox.showerror("eBay Error", str(e))

    def on_ebay_list(self):
        card_id = self._get_selected_id()
        if card_id is None:
            return

        cfg = _load_ebay_config()
        if not cfg.get("access_token"):
            messagebox.showwarning("eBay", "Not connected to eBay.\nGo to the Settings tab first.")
            return

        card = load_card(card_id)
        if card is None:
            return
        total_qty = int(card.get("quantity", 1))

        win = tk.Toplevel(self.root)
        win.title("List on eBay")
        win.resizable(False, False)
        win.grab_set()

        frame = ttk.Frame(win, padding=15)
        frame.pack()

        ttk.Label(frame, text=f"{card['name']} — {card['set']}",
                  font=("Helvetica", 12, "bold")).grid(row=0, column=0, columnspan=2, pady=(0, 10))

        ttk.Label(frame, text="Price (EUR):").grid(row=1, column=0, sticky="e", padx=(0, 10), pady=4)
        price_entry = ttk.Entry(frame, width=15)
        price_entry.insert(0, f"{card['price']:.2f}" if card["price"] else "")
        price_entry.grid(row=1, column=1, sticky="w", pady=4)

        ttk.Label(frame, text="Quantity:").grid(row=2, column=0, sticky="e", padx=(0, 10), pady=4)
        qty_entry = ttk.Entry(frame, width=8)
        qty_entry.insert(0, str(total_qty))
        qty_entry.grid(row=2, column=1, sticky="w", pady=4)

        status_lbl = ttk.Label(frame, text="")
        status_lbl.grid(row=5, column=0, columnspan=2, pady=(5, 0))

        def _validate_input():
            try:
                price = float(price_entry.get().strip().replace(",", "."))
                qty = int(qty_entry.get().strip())
            except ValueError:
                messagebox.showwarning("Invalid Input", "Price must be a number, quantity a whole number.")
                return None, None
            if qty < 1 or qty > total_qty:
                messagebox.showwarning("Invalid Quantity", f"Must be between 1 and {total_qty}.")
                return None, None
            return price, qty

        def save_draft():
            price, qty = _validate_input()
            if price is None:
                return
            status_lbl.config(text="Creating draft...", foreground="orange")
            win.update()
            try:
                offer_id, _ = ebay_create_draft(card_id, price, qty)
                win.destroy()
                messagebox.showinfo("eBay", f"Draft saved!\n\nOffer ID: {offer_id}\n\n"
                                    "The draft is stored in the eBay API.\n"
                                    "Use 'List on eBay' to publish later.")
            except Exception as e:
                status_lbl.config(text="Failed", foreground="red")
                messagebox.showerror("eBay Error", str(e))

        def publish_now():
            price, qty = _validate_input()
            if price is None:
                return
            status_lbl.config(text="Publishing listing...", foreground="orange")
            win.update()
            try:
                offer_id, _ = ebay_create_draft(card_id, price, qty)
                listing_id = ebay_publish_offer(offer_id)
                win.destroy()
                messagebox.showinfo("eBay", f"Listing published!\n\nListing ID: {listing_id}\n\n"
                                    f"View at: ebay.de/itm/{listing_id}")
            except Exception as e:
                status_lbl.config(text="Failed", foreground="red")
                messagebox.showerror("eBay Error", str(e))

        btn_row = ttk.Frame(frame)
        btn_row.grid(row=3, column=0, columnspan=2, pady=(10, 0))
        ttk.Button(btn_row, text="Save as Draft", command=save_draft).pack(side="left", padx=5)
        ttk.Button(btn_row, text="Publish Now", command=publish_now).pack(side="left", padx=5)
        price_entry.focus()

    def _on_tree_double_click(self, event):
        item = self.tree.identify_row(event.y)
        if not item:
            return
        card_id = int(item)
        col = self.tree.identify_column(event.x)
        if col == "#2":
            photo_path = get_photo_path(card_id)
            if photo_path:
                self._show_photo_popup(card_id, photo_path)
            return
        url = get_card_url(card_id)
        if url:
            webbrowser.open(url)
        else:
            messagebox.showinfo("No Link", f"No Cardmarket link saved for card #{card_id}.")

    def on_view_photo(self):
        card_id = self._get_selected_id()
        if card_id is None:
            return
        photo_path = get_photo_path(card_id)
        if not photo_path:
            messagebox.showinfo("No Photo", f"No photo for card #{card_id}.")
            return
        self._show_photo_popup(card_id, photo_path)

    def _show_photo_popup(self, card_id, photo_path):
        img = Image.open(photo_path)
        from PIL import ImageOps
        img = ImageOps.exif_transpose(img)
        max_w, max_h = 600, 700
        img.thumbnail((max_w, max_h), Image.LANCZOS)
        win = tk.Toplevel(self.root)
        win.title(f"Photo #{card_id:03d}")
        win.resizable(False, False)
        tk_img = ImageTk.PhotoImage(img)
        label = ttk.Label(win, image=tk_img)
        label.image = tk_img
        label.pack(padx=5, pady=5)
        win.bind("<Escape>", lambda e: win.destroy())
        win.bind("<space>", lambda e: win.destroy())

    def on_delete_photo(self):
        card_id = self._get_selected_id()
        if card_id is None:
            return
        photo_path = get_photo_path(card_id)
        if not photo_path:
            messagebox.showinfo("No Photo", f"Card #{card_id} has no photo.")
            return
        if messagebox.askyesno("Delete Photo", f"Delete photo for card #{card_id}?"):
            os.remove(photo_path)
            self.refresh_collection()

    def on_add_photo(self):
        card_id = self._get_selected_id()
        if card_id is None:
            return
        path = filedialog.askopenfilename(
            title=f"Select photo for card #{card_id}",
            filetypes=[("Images", "*.jpg *.jpeg *.png *.heic *.heif")],
        )
        if path:
            save_photo(card_id, path)
            self.refresh_collection()

    def on_delete(self):
        card_id = self._get_selected_id()
        if card_id is None:
            return
        if not messagebox.askyesno("Delete Card", f"Delete card #{card_id}? This cannot be undone."):
            return
        delete_card(card_id)
        self.refresh_collection()

    def on_check_update(self):
        try:
            has_update, remote_version, content = check_for_update()
        except Exception as e:
            messagebox.showerror("Update Error", f"Could not check for updates:\n{e}")
            return

        if not has_update:
            messagebox.showinfo("Up to Date", f"You are running the latest version (v{VERSION}).")
            return

        if not messagebox.askyesno(
            "Update Available",
            f"New version v{remote_version} available.\n"
            f"Current version: v{VERSION}\n\n"
            "Download and install?\n"
            "(Your data will not be affected.)",
        ):
            return

        try:
            apply_update(content)
            messagebox.showinfo("Updated", f"Updated to v{remote_version}.\nThe app will restart now.")
            self._save_state()
            restart_app()
        except Exception as e:
            messagebox.showerror("Update Error", f"Could not install update:\n{e}")


def _activate_window():
    import subprocess
    subprocess.Popen([
        "osascript", "-e",
        'tell application "System Events" to set frontmost of first process '
        'whose unix id is ' + str(os.getpid()) + ' to true'
    ], stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)


if __name__ == "__main__":
    photo_server = start_photo_server()

    root = tk.Tk()
    app = App(root)
    root.after(300, _activate_window)
    root.mainloop()
