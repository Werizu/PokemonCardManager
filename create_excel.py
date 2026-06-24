#!/usr/bin/env python3
"""Creates an empty Pokemon-Inventar.xlsx with Inventar, Sales, and Dashboard sheets."""
import sys
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import CellIsRule
from openpyxl.worksheet.datavalidation import DataValidation

def create_workbook(path):
    wb = Workbook()

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    # ── Inventar sheet ──
    ws = wb.active
    ws.title = "Inventar"
    inv_headers = [
        ("ID", 6), ("Card Name", 18), ("Set / Edition", 24),
        ("Card Number", 14), ("Language", 12), ("Condition (Raw)", 14),
        ("Grading Status", 16), ("Grading Service", 15), ("Grade", 10),
        ("Cert Number", 15), ("Purchase Price", 16), ("Purchase Date", 14),
        ("Source", 13), ("Quantity", 10), ("Status", 14),
        ("Grading Vormerkung", 16), ("Vormerkung Notiz", 22),
        ("Vorheriger Service", 15), ("Vorheriger Grade", 14), ("Vorheriger Cert#", 15),
    ]
    for col, (name, width) in enumerate(inv_headers, 1):
        cell = ws.cell(row=1, column=col, value=name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
        ws.column_dimensions[cell.column_letter].width = width

    ws.freeze_panes = "A2"

    dv_lang = DataValidation(type="list", formula1='"EN,DE,JP,KR,FR,IT,ES,CN,PT"', allow_blank=True)
    dv_cond = DataValidation(type="list", formula1='"Mint,Near Mint,Excellent,Good,Light Played,Played,Poor"', allow_blank=True)
    dv_gs = DataValidation(type="list", formula1='"Not Graded,Submitted,Graded"', allow_blank=True)
    dv_svc = DataValidation(type="list", formula1='"PSA,CGC,Beckett,ACE,TAG"', allow_blank=True)
    dv_src = DataValidation(type="list", formula1='"Cardmarket,eBay,Convention,Trade,Booster,Other"', allow_blank=True)
    dv_status = DataValidation(type="list", formula1='"Collection,For Sale,Sold"', allow_blank=True)
    dv_vormerkung = DataValidation(type="list", formula1='"Ja"', allow_blank=True)

    dv_lang.sqref = "E2:E200"
    dv_cond.sqref = "F2:F200"
    dv_gs.sqref = "G2:G200"
    dv_svc.sqref = "H2:H200"
    dv_src.sqref = "M2:M200"
    dv_status.sqref = "O2:O200"
    dv_vormerkung.sqref = "P2:P200"

    for dv in [dv_lang, dv_cond, dv_gs, dv_svc, dv_src, dv_status, dv_vormerkung]:
        ws.add_data_validation(dv)

    # ── Sales sheet ──
    ws_s = wb.create_sheet("Sales")
    sales_headers = [
        ("ID", 6), ("Card Name", 18), ("Sale Date", 14),
        ("Platform", 13), ("Sale Price", 13), ("Fees", 12),
        ("Shipping Cost", 14), ("Net Revenue", 13),
        ("Purchase Price", 13), ("Net Profit", 12),
        ("Buyer", 16), ("Tracking Number", 18),
    ]
    for col, (name, width) in enumerate(sales_headers, 1):
        cell = ws_s.cell(row=1, column=col, value=name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
        ws_s.column_dimensions[cell.column_letter].width = width

    ws_s.freeze_panes = "A2"

    dv_plat = DataValidation(type="list", formula1='"Cardmarket,eBay,Trade"', allow_blank=True)
    dv_plat.sqref = "D2:D200"
    ws_s.add_data_validation(dv_plat)

    # Sales formulas (rows 2-200)
    for r in range(2, 201):
        ws_s.cell(row=r, column=8).value = f'=IF(E{r}<>"",E{r}-F{r}-G{r},"")'
        ws_s.cell(row=r, column=8).number_format = '#,##0.00 "EUR"'
        ws_s.cell(row=r, column=10).value = f'=IF(AND(H{r}<>"",I{r}<>""),H{r}-I{r},"")'
        ws_s.cell(row=r, column=10).number_format = '#,##0.00 "EUR"'

    # ── Dashboard sheet ──
    ws_d = wb.create_sheet("Dashboard")
    ws_d.column_dimensions["A"].width = 4
    ws_d.column_dimensions["B"].width = 34
    ws_d.column_dimensions["C"].width = 22
    ws_d.column_dimensions["D"].width = 4
    ws_d.column_dimensions["E"].width = 22
    ws_d.column_dimensions["F"].width = 13

    title_font = Font(bold=True, size=16, color="4472C4")
    section_font = Font(bold=True, size=13, color="4472C4")
    label_font = Font(size=11)
    value_font = Font(bold=True, size=11)

    ws_d.cell(row=2, column=2, value="Pokemon Collection - Dashboard").font = title_font
    ws_d.cell(row=4, column=2, value="Overview").font = section_font

    labels = [
        (5, "Total Cards", '=COUNTA(Inventar!A2:A200)'),
        (6, "Raw", '=COUNTIF(Inventar!G2:G200,"Not Graded")'),
        (7, "Submitted for Grading", '=COUNTIF(Inventar!G2:G200,"Submitted")'),
        (8, "Graded", '=COUNTIF(Inventar!G2:G200,"Graded")'),
        (10, "Total Invested", '=SUMPRODUCT((Inventar!K2:K200<>"")*(Inventar!M2:M200<>"Trade"),Inventar!K2:K200)'),
    ]
    for row, label, formula in labels:
        ws_d.cell(row=row, column=2, value=label).font = label_font
        c = ws_d.cell(row=row, column=3, value=formula)
        c.font = value_font
        if "Invest" in label:
            c.number_format = '#,##0.00 "EUR"'

    ws_d.cell(row=12, column=2, value="Sales").font = section_font

    sales_labels = [
        (13, "Total Sales", '=COUNTA(Sales!A2:A200)'),
        (14, "Total Net Revenue", '=SUMPRODUCT((Sales!H2:H200<>"")*(Sales!D2:D200<>"Trade"),Sales!H2:H200)'),
        (15, "Total Net Profit", '=SUMPRODUCT((Sales!J2:J200<>"")*(Sales!D2:D200<>"Trade"),Sales!J2:J200)'),
        (16, "Total Fees Paid", '=SUMPRODUCT((Sales!F2:F200<>"")*(Sales!D2:D200<>"Trade"),Sales!F2:F200)'),
    ]
    for row, label, formula in sales_labels:
        ws_d.cell(row=row, column=2, value=label).font = label_font
        c = ws_d.cell(row=row, column=3, value=formula)
        c.font = value_font
        if row > 13:
            c.number_format = '#,##0.00 "EUR"'

    wb.save(path)

if __name__ == "__main__":
    create_workbook(sys.argv[1])
